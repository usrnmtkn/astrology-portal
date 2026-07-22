#!/usr/bin/env python3
"""Import edited wording from admin/source-wording-review.xlsx back into
source-rows/fallback-source-rows-v3.json, then re-run the package test.

Usage: python3 admin/import-wording-updates.py [path-to-edited.xlsx]

Rules applied on import:
- Blank NEW wording  -> row unchanged.
- NEW wording text   -> body replaced verbatim, review_status set to 'approved'
                        (a human wrote it), original body preserved in note.
- NEW wording DELETE -> row removed from the vocabulary (slot suppresses).
- Em dashes and the contract's bannedWords are rejected with an error.
"""
import json, re, subprocess, sys, os

here = os.path.dirname(os.path.abspath(__file__))
pkg = os.path.dirname(here)
xlsx = sys.argv[1] if len(sys.argv) > 1 else os.path.join(here, "source-wording-review.xlsx")
rows_path = os.path.join(pkg, "source-rows", "fallback-source-rows-v3.json")
contract = json.load(open(os.path.join(pkg, "contracts", "CONTENT-ROLE-CONTRACT.json")))
banned = [w.lower() for w in contract.get("styleRules", {}).get("bannedWords", [])]

from openpyxl import load_workbook
ws = load_workbook(xlsx)["Source wording"]
edits, deletes, errors = {}, set(), []
for row in ws.iter_rows(min_row=2, values_only=True):
    key, new = row[1], row[5]
    note = (row[7] or "")
    if not key or "EXAMPLE ROW" in str(note): continue
    if new is None or not str(new).strip(): continue
    new = str(new).strip()
    if new.upper() == "DELETE": deletes.add(key); continue
    if "—" in new or "–" in new: errors.append(f"{key}: em/en dash not allowed"); continue
    low = new.lower()
    for w in banned:
        if re.search(rf"\b{re.escape(w)}\b", low): errors.append(f"{key}: banned word '{w}'")
    edits[key] = new

if errors:
    print("REJECTED, fix these first:"); [print(" -", e) for e in errors]; sys.exit(1)

d = json.load(open(rows_path))
n_e = n_d = 0
kept = []
for r in d["vocabularyRows"]:
    k = r["contentKey"]
    if k in deletes: n_d += 1; continue
    if k in edits:
        r["note"] = (r.get("note", "") + f" | replaced via wording review; was: {r['body']}").strip(" |")
        r["body"] = edits[k]; r["review_status"] = "approved"; n_e += 1
    kept.append(r)
d["vocabularyRows"] = kept
json.dump(d, open(rows_path, "w"), indent=1, ensure_ascii=False)
print(f"Applied {n_e} edits, {n_d} deletions.")
res = subprocess.run(["node", os.path.join(pkg, "tests", "verify-fallback-architecture.mjs")])
sys.exit(res.returncode)
